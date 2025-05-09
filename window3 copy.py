import os
from openpyxl import load_workbook


# code framework for window 3
accountID = None
accountType = None
firstName = None
lastName = None
email = None
contact_number = None
nationality = None
religion = None
sex = None
civil_status = None
age = None
isDisability = None
permanent_address = None
password = None

def load_account_data(account_id):
    global accountID, accountType, firstName, lastName, email, contact_number, nationality, religion, sex, civil_status, age, isDisability, permanent_address, password

    try:
        wb = load_workbook("AccountDatabase.xlsx")
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            current_id = str(row[0]).strip()
            if current_id == account_id:
                print(f"Details for Account ID: {account_id}")
                for header, value in zip(headers, row):
                    
                    if header == "ID Number":
                        accountID = value
                        print(f"Account ID: {accountID}")
                    elif header == "Account Type":
                        accountType = value
                        print(f"Account Type: {accountType}")
                    elif header == "First Name":
                        firstName = value
                        print(f"Account First Name: {firstName}")
                    elif header == "Last Name":
                        lastName = value
                        print(f"Account Last Name: {lastName}")
                    elif header == "Email":
                        email = value
                        print(f"Account Email: {email}")
                    elif header == "Contact Number":
                        contact_number = value
                        print(f"Contact Number: {contact_number}")
                    elif header == "Nationality":
                        nationality = value
                        print(f"Nationality : {nationality}")
                    elif header == "Religion":
                        religion = value
                        print(f"Religion : {religion}")
                    elif header == "Sex":
                        sex = value
                        print(f"Sex : {sex}")
                    elif header == "Civil Status":
                        civil_status = value
                        print(f"Civil Status : {civil_status}")
                    elif header == "Age":
                        age = value
                        print(f"Age : {age}")
                    elif header == "Disability":
                        isDisability = value
                        print(f"Disabillity : {isDisability}")
                    elif header == "Permanent Address":
                        permanent_address = value
                        print(f"Permanent Address : {permanent_address}")
                    elif header == "Password":
                        password = value
                        print(f"Password : {password}")
                    
                    # print(f"{header}: {value}")
                return accountID, accountType, firstName, lastName, email, contact_number, nationality, religion, sex, civil_status, age, isDisability, permanent_address, password

        print(f"Account ID {account_id} not found.")

    except FileNotFoundError:
        print("ERROR: AccountDatabase.xlsx not found.")

if __name__ == "__main__":
    account_id = os.environ.get("ACCOUNT_ID")
    
    if not account_id:
        print("ERROR: ACCOUNT_ID environment variable not set.")
    else:
        load_account_data(account_id)
        print(accountID, accountType, lastName, email, contact_number, nationality, religion, sex, civil_status, age, isDisability, permanent_address, password)
