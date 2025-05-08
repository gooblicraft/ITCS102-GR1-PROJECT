import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import openpyxl

# def search_data():
#     input_name = name_entry.get().strip()
    
#     try:
#         wb = load_workbook('AccountDatabase.xlsx')  # Replace with your Excel file name
#         ws = wb.active

#         for row in ws.iter_rows(min_row=2, values_only=True):
#             if row[0] == input_name:
#                 name_value.set(row[2])
#                 ID_value.set(row[1])
#                 pass_value.set(row[2])
#                 return
        
#         # If not found
#         messagebox.showinfo("Not Found", f"'{input_name}' not found in the Excel file.")
#         name_value.set("")
#         ID_value.set("")
#         pass_value.set("")

#     except FileNotFoundError:
#         messagebox.showerror("Error", "Excel file not found. Please check the file name/path.")

# # GUI Setup
# root = tk.Tk()
# root.title("Excel Lookup")

# tk.Label(root, text="Enter Name:").grid(row=0, column=0, padx=5, pady=5)
# name_entry = tk.Entry(root)
# name_entry.grid(row=0, column=1, padx=5, pady=5)

# tk.Button(root, text="Search", command=search_data).grid(row=0, column=2, padx=5, pady=5)

# name_value = tk.StringVar()
# ID_value = tk.StringVar()
# pass_value = tk.StringVar()

# tk.Label(root, text="Name:").grid(row=1, column=0, padx=5, pady=5)
# tk.Entry(root, textvariable=name_value, state='readonly').grid(row=1, column=1, padx=5, pady=5)

# tk.Label(root, text="ID:").grid(row=2, column=0, padx=5, pady=5)
# tk.Entry(root, textvariable=ID_value, state='readonly').grid(row=2, column=1, padx=5, pady=5)

# tk.Label(root, text="Pass:").grid(row=3, column=0, padx=5, pady=5)
# tk.Entry(root, textvariable=pass_value, state='readonly').grid(row=3, column=1, padx=5, pady=5)

# root.mainloop()


workbook = openpyxl.load_workbook("AccountDatabase.xlsx")
worksheet = workbook['Sheet']
columns = worksheet.iter_cols()

for column in columns:
    print(column)