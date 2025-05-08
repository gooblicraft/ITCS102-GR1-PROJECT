import openpyxl
import os
from tkinter import *

# ========== INITIALIZE EXCEL ==========
def initialize_excel_file():
    file_name = "karlo.xlsx"
    sheet_name = "Accounts"
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Student ID", "Password", "Name"])  # Header row
        wb.save(file_name)
    else:
        wb = openpyxl.load_workbook(file_name)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(["Student ID", "Password", "Name"])
            wb.save(file_name)
        else:
            ws = wb[sheet_name]
            if ws.max_row < 1:
                ws.append(["Student ID", "Password", "Name"])
                wb.save(file_name)

initialize_excel_file()

# Show/Hide password logic for main login
password_visible = False

def toggle_password_button():
    global password_visible
    if entry_2.get() != "Password":
        if password_visible:
            entry_2.config(show="*")
            show_button.config(text="Show")
            password_visible = False
        else:
            entry_2.config(show="")
            show_button.config(text="Hide")
            password_visible = True

# Placeholder Handling Functions
def on_entry_click(entry, placeholder, is_password=False):
    if entry.get() == placeholder:
        entry.delete(0, "end")
        entry.config(fg="gray33")
        if is_password:
            if not password_visible:
                entry.config(show="*")
            else:
                entry.config(show="")

def on_focusout(entry, placeholder, is_password=False):
    if entry.get() == "":
        entry.insert(0, placeholder)
        entry.config(fg="#767676")
        if is_password:
            entry.config(show="")

# ========== LOGIN WITH EXCEL ==========
def check_credentials(student_id, password):
    try:
        wb = openpyxl.load_workbook("karlo.xlsx")
        sheet = wb["Accounts"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            stored_id = row[0]
            stored_password = row[1]
            if student_id == stored_id and password == stored_password:
                return row[2]
        return None
    except Exception as e:
        print(f"Error reading file: {e}")
        return None

def handle_login():
    student_id = entry_1.get()
    password = entry_2.get()

    if student_id == "Student ID" or student_id.strip() == "":
        if password == "Password" or password.strip() == "":
            warning_label.config(text="Please fill up both forms.")
        else:
            warning_label.config(text="Please fill up the Student ID form.")
    elif password == "Password" or password.strip() == "":
        warning_label.config(text="Please fill up the Password form.")
    else:
        warning_label.config(text="")
        name = check_credentials(student_id, password)
        if name:
            warning_label.config(text=f"Welcome, {name}!")
        else:
            warning_label.config(text="Invalid Student ID or Password.")

# ========== CREATE ACCOUNT WINDOW ==========
def open_create_account():
    create_accW = Toplevel(window)
    create_accW.title("Create Account")
    create_accW.geometry("500x500")
    create_accW.config(bg="white")

    Label(create_accW, text="Create New Account", font=("JetsBrains Mono", 14, "bold"), bg="white").pack(pady=20)

    Label(create_accW, text="Username", font=("JetsBrains Mono", 10), bg="white").pack()
    username_entry = Entry(create_accW)
    username_entry.pack(pady=5)

    Label(create_accW, text="Password", font=("JetsBrains Mono", 10), bg="white").pack()
    password_entry = Entry(create_accW, show="*")
    password_entry.pack(pady=5)

    def toggle_create_password():
        if password_entry.cget('show') == '':
            password_entry.config(show='*')
            show_pw_btn.config(text='Show')
        else:
            password_entry.config(show='')
            show_pw_btn.config(text='Hide')

    show_pw_btn = Button(create_accW, text="Show", command=toggle_create_password,
                         font=("JetBrains Mono", 8), bg="white", bd=0, cursor="hand2")
    show_pw_btn.pack(pady=2)

    def register_and_close():
        username = username_entry.get()
        password = password_entry.get()
        try:
            wb = openpyxl.load_workbook("karlo.xlsx")
            sheet = wb["Accounts"]
            sheet.append([username, password, username])
            wb.save("karlo.xlsx")
            print("Account Created:", username)
            create_accW.destroy()
        except Exception as e:
            print(f"Error saving account: {e}")
            warning_label.config(text="Error creating account.")

    Button(create_accW, text="Register", font=("JetsBrains Mono", 10), command=register_and_close).pack(pady=20)

# ========== MAIN WINDOW ==========
window = Tk()
window.geometry("665x410")
window.configure(bg="#FFFFFF")
window.title("Window 1")

canvas = Canvas(
    window,
    bg="#FFFFFF",
    height=410,
    width=665,
    bd=0,
    highlightthickness=0,
    relief="ridge"
)
canvas.place(x=0, y=0)

# ========== IMAGES ==========
image_image_1 = PhotoImage(file="beta 0.1\\assets\\frame0\\image_1.png")
image_1 = canvas.create_image(519.0, 126.0, image=image_image_1)

image_image_2 = PhotoImage(file="beta 0.1\\assets\\frame0\\image_2.png")
image_2 = canvas.create_image(519.0, 190.0, image=image_image_2)

image_image_3 = PhotoImage(file="beta 0.1\\assets\\frame0\\image_3.png")
image_3 = canvas.create_image(140.0, 180.0, image=image_image_3)

# ========== LABEL TEXTS ==========
canvas.create_text(450.0, 286.0, anchor="nw", text="          Forgot Password?", fill="#757575", font=("JetBrains Mono", 10 * -1))
canvas.create_text(466.0, 76.0, anchor="nw", text="Sign in your account", fill="#757575", font=("JetBrains Mono", 10 * -1))
canvas.create_text(419.0, 38.0, anchor="nw", text="ScanTracker", fill="#060606", font=("JetBrains Mono", 24, "bold"))

# ========== ENTRY ==========
entry_1 = Entry(
    bd=0,
    bg="#AEAEAE",
    fg="#767676",
    highlightthickness=0,
    font=("JetBrains Mono", 10, "bold")
)
entry_1.insert(0, "Student ID")
entry_1.place(x=464.0, y=118.0, width=152.0, height=14.0)
entry_1.bind("<FocusIn>", lambda event: on_entry_click(entry_1, "Student ID"))
entry_1.bind("<FocusOut>", lambda event: on_focusout(entry_1, "Student ID"))

entry_2 = Entry(
    bd=0,
    bg="#AEAEAE",
    fg="#767676",
    highlightthickness=0,
    font=("JetBrains Mono", 10, "bold")
)
entry_2.insert(0, "Password")
entry_2.place(x=464.0, y=182.0, width=152.0, height=14.0)
entry_2.bind("<FocusIn>", lambda event: on_entry_click(entry_2, "Password", is_password=True))
entry_2.bind("<FocusOut>", lambda event: on_focusout(entry_2, "Password", is_password=True))

# ========== WARNING LABEL ==========
warning_label = Label(window, text="", fg="red", bg="#FFFFFF", font=("JetBrains Mono", 9, "bold"))
warning_label.place(x=402, y=290)

# ========== BUTTONS ==========
button_image_1 = PhotoImage(file="beta 0.1\\assets\\frame0\\button_1.png")
button_logIn = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=handle_login,
    relief="flat",
    cursor="hand2"
)
button_logIn.place(x=402.0, y=224.0, width=234.0, height=59.0)

show_button = Button(
    window,
    text="Show",
    command=toggle_password_button,
    font=("JetBrains Mono", 7),
    bd=0,
    bg="#AEAEAE",
    activebackground="#AEAEAE",
    fg="gray25",
    cursor="hand2"
)
show_button.place(x=568, y=180, width=40, height=18)

button_image_2 = PhotoImage(file="beta 0.1\\assets\\frame0\\button_2.png")
button_2 = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    command=open_create_account,
    relief="flat",
    cursor="hand2"
)
button_2.place(x=402.0, y=325.0, width=234.0, height=57.0)

window.resizable(False, False)
window.mainloop()
