from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import os
from openpyxl import *
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter    

def inClick():
    createExcel()
    submit_data()

#<===================================================== OPENPYXL===================================================>

filename = "AccountDatabase.xlsx"

def createExcel():
    if not os.path.exists(filename):
        workbook = Workbook()
        ws = workbook.active
        ws.append(["ID Number","Account Type","First Name", "Last Name", 
                   "Email","Contact Number", "Nationality", "Religion", 
                   "Sex", "Civil Status", "Age", "Disability", "Permanent Address","Password"])  
        workbook.save(filename)

def submit_data():

    path = filename
    workbook = load_workbook(path)
    sheet = workbook.active

    first_name = Fname_entry.get()
    last_name = Lname_entry.get()
    email = email_entry.get()
    contact_num = contact_entry.get()
    nationality = nationality_entry.get()
    religion = cb_religion.get()
    sex = cb_sex.get()
    civil_status = cb_civil_status.get()
    age = age_entry.get()
    disability = disability_RB.get()
    permanent_address = address_entry.get()
    password = real_pass
    
    row_values = [first_name, last_name, email, contact_num, nationality, religion, sex, civil_status, age, disability, permanent_address, password]
    sheet.append(row_values)
    workbook.save(path)















real_pass = None
# real password (TRIGGER WINDOW SWITCH)
def get_pass():
    pass1 = confirm_pass.get()
    pass2 = set_pass.get()


    if pass1 == pass2:
        real_pass = pass1
        print(real_pass)
        return real_pass
    else:
        messagebox.showerror("Error", "Your password doesn't match")

window = Tk()
window.title("Window 2 Create Account")
window.geometry("670x410")
window.configure(bg = "#FFFFFF")

# Styling ttk 
style = ttk.Style()
style.theme_use('clam')

# Define custom layout without border
style.layout('Custom.TCombobox',
            [('Combobox.downarrow', {'side': 'right', 'sticky': ''}),
            ('Combobox.padding', {'expand': '1', 'sticky': 'nswe',
            'children': [('Combobox.textarea', {'sticky': 'nswe'})]})])

# Configure background and remove border
style.configure('Custom.TCombobox',
                fieldbackground='#ffffff',  # Entry field background
                background='#ffffff',       # Dropdown background
                borderwidth=0,
                relief='flat',
                padding=0)

canvas = Canvas(
    window,
    bg = "#FFFFFF",
    height = 410,
    width = 665,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge", 
    scrollregion=(0, 0, 670, 550)
)

canvas.place(x = 0, y = 0)
y_scroll = Scrollbar(window, orient="vertical", command=canvas.yview)
canvas.config(yscrollcommand=y_scroll.set)
y_scroll.place(x=651, y=0, height=410)

canvas.create_text(
    76.0,
    22.0,
    anchor="nw",
    text="// Our Logo here",
    fill="#000716",
    font=("JetBrainsMono Bold", 12 * -1)
)

# ======== IMAGES =========

image_image_1 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_1.png")
image_1 = canvas.create_image(
    444.0,
    33.0,
    image=image_image_1
)

image_image_2 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_2.png")
image_2 = canvas.create_image(
    276.0,
    180.0,
    image=image_image_2
)

image_image_4 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_4.png")
image_4 = canvas.create_image(
    261.0,
    237.0,
    image=image_image_4
)

image_image_5 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_5.png")
image_5 = canvas.create_image(
    485.0,
    180.0,
    image=image_image_5
)

image_image_7 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_7.png")
image_7 = canvas.create_image(
    538.0,
    237.0,
    image=image_image_7
)

image_image_8 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_8.png")
image_8 = canvas.create_image(
    39.0,
    33.0,
    image=image_image_8
)

image_image_9 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_9.png")
image_9 = canvas.create_image(
    436.0,
    110.0,
    image=image_image_9
)
# Image Entry FIRST NAME
image_image_10 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_10.png")
image_10 = canvas.create_image(
    333.0,
    203.0,
    image=image_image_10
)

image_image_11 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_11.png")
image_11 = canvas.create_image(
    548.0,
    203.0,
    image=image_image_11
)

# ========  Set Password ========= 
# Entry set password
image_image_3 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_3.png")
image_3 = canvas.create_image(
    280.0,
    450.0,
    image=image_image_3
)
image_image_12 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_12.png")

# image entry SET PASSWORD
image_12 = canvas.create_image(
    330.0,
    470.0,
    image=image_image_12
)

# ============ Confirm password ===========

# image confirm password
image_image_6 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_6.png")
image_6 = canvas.create_image(
    500.0,
    450.0,
    image=image_image_6
)

# Border confirm password
image_image_14 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_14.png")
image_14 = canvas.create_image(
    540.0,
    470.0,
    image=image_image_14
)

# ======== Whole Mid code =====
# Image MID
image_image_17 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_17.png")
image_17 = canvas.create_image(
    434.0,
    360.0,
    image=image_image_17
)

# ==================== COMBOBOX AND RADIO BUTTON =========================

# Combobox religion
cb_religion = ttk.Combobox(values=["Catholic", "INC", "Muslim"], style='Custom.TCombobox', width=25)
cb_religion.set("Select")
canvas.create_window(540, 316, window=cb_religion)

# Combobox Sex
cb_sex = ttk.Combobox(values=["Male", "Female"], style='Custom.TCombobox', width=20)
cb_sex.set("Select")
canvas.create_window(310, 370, window=cb_sex)

# Combobox Civil Satus
cb_civil_status = ttk.Combobox(values=["Single", "Married"], style='Custom.TCombobox', width=20)
cb_civil_status.set("Select")
canvas.create_window(470, 370, window=cb_civil_status)

# Disability Radio Button
disability_RB = StringVar()
rb_yes_disable = Radiobutton(
    window, 
    text="yes", 
    variable=disability_RB, 
    value=1,
    bg="#ffffff", 
    activebackground="#ffffff", 
    highlightthickness=0, 
    bd=0)
canvas.create_window(250, 422, window=rb_yes_disable)

rb_no_disable = Radiobutton(
    window, 
    text="no", 
    variable=disability_RB, 
    value=0,
    bg="#ffffff", 
    activebackground="#ffffff",
    highlightthickness=0, 
    bd=0)
canvas.create_window(300, 422, window=rb_no_disable)

# ====================== ENTRY ======================

# First Name entry
Fname_entry = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#767676",
    font= ("JetBrains Mono", 10 * -1),
    highlightthickness=0
)
canvas.create_window(334, 205, window=Fname_entry, width=165, height=12)

# Lastname entry
Lname_entry = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#767676",
    font= ("JetBrains Mono", 10 * -1),
    highlightthickness=0
)
canvas.create_window(551, 204, window=Lname_entry, width=165, height=12)

# Email Entry
email_entry = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#767676",
    font= ("JetBrains Mono", 10 * -1),
    highlightthickness=0
)
canvas.create_window(350, 260, window=email_entry, width=212, height=12)

# Contact Number entry
contact_entry = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#767676",
    font= ("JetBrains Mono", 10 * -1),
    highlightthickness=0
)
canvas.create_window(570, 260, window=contact_entry, width=122, height=12)

# Entry nationality
address_entry = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#767676",
    font= ("JetBrains Mono", 10 * -1),
    highlightthickness=0
)
canvas.create_window(480, 422, window=address_entry, width=276, height=12)

# Entry Age
age_entry = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#767676",
    font= ("JetBrains Mono", 10 * -1),
    highlightthickness=0
)
canvas.create_window(595, 370, window=age_entry, width=60, height=12)

# Entry nationality
nationality_entry = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#767676",
    font= ("JetBrains Mono", 10 * -1),
    highlightthickness=0
)
canvas.create_window(325, 316, window=nationality_entry, width=160, height=12)

# entry confirm password
confirm_pass = Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#767676",
    font= ("JetBrains Mono", 10 * -1),
    highlightthickness=0
)
canvas.create_window(540, 470, window=confirm_pass, width=165, height=12)

# entry SET PASSOWRD
set_pass= Entry(
    bd=0,
    bg="#FFFFFF",
    fg="#767676",
    font= ("JetBrains Mono", 10 * -1),
    highlightthickness=0
)
canvas.create_window(330, 470, window=set_pass, width=165, height=12)

image_image_16 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_16.png")
image_16 = canvas.create_image(
    108.0,
    216.0,
    image=image_image_16
)

# ============== BUTTONS ===============

# Button Facilitator
button_image_2 = PhotoImage(file="beta 0.2\\assets\\frame0\\button_2.png")
button_facilitator = Button(
    image=button_image_2,
    borderwidth=0,
    highlightthickness=0,
    command= inClick,
    relief="flat"
)
canvas.create_window(330, 520, window=button_facilitator, width=177, height=45)


# Button Attendee
button_image_3 = PhotoImage(file="beta 0.2\\assets\\frame0\\button_3.png")
button_attendee = Button(
    image=button_image_3,
    borderwidth=0,
    border=0,
    highlightthickness=0,
    command= inClick,
    relief="flat"
)
canvas.create_window(540, 520, window=button_attendee, width=177, height=45)

# Button log im
button_image_1 = PhotoImage(file="beta 0.2\\assets\\frame0\\button_1.png")
button_logIn = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=get_pass,
    relief="flat"
)
canvas.create_window(595, 32, window=button_logIn, width=103, height=31)

image_image_13 =PhotoImage(file="beta 0.2\\assets\\frame0\\image_13.png")
image_13 = canvas.create_image(
    352.0,
    260.0,
    image=image_image_13
)

image_image_15 = PhotoImage(file="beta 0.2\\assets\\frame0\\image_15.png")
image_15 = canvas.create_image(
    570.0,
    263.0,
    image=image_image_15
)


window.resizable(False, False)
window.mainloop()
