
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import *


window = Tk()
window.title("Window 2 Create Account")
window.geometry("665x410")
window.configure(bg = "#FFFFFF")

file_name = "Data.xlsx"


def excelll():
    try:
        return load_workbook("StudentScoreRecord.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["First_name","Last_name","Email","password"])
        wb.save(file_name)
        return wb





excelll



window.mainloop()