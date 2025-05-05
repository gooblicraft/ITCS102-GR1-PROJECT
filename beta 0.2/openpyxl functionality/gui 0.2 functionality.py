
from tkinter import messagebox
from openpyxl import Workbook , load_workbook
import tkinter as tk
from tkinter import *


window = Tk()
window.title("Window 2 Create Account")
window.geometry("665x410")
window.configure(bg = "#FFFFFF")

file_name = "Data.xlsx"


def excelll():
    


    wb = Workbook()
    ws = wb.active
    ws.append(["First_name","Last_name","Email","password"])
    wb.save(file_name)
    
def save():

    fn = entry1.get()
    ln = entry2.get()
    em = entry3.get()
    pas = entry4.get()
    wb = load_workbook(file_name)
    ws = wb.active
    ws.append([fn,ln,em,pas])
    wb.save(file_name)


excelll()

entry1 = Entry(window)
entry1.pack()
entry2 = Entry(window)
entry2.pack()
entry3 = Entry(window)
entry3.pack()
entry4 = Entry(window)
entry4.pack()
entryy4 = Entry(window)
entryy4.pack()
button = Button(window, text="Save", command=save)
button.pack()

if entry4 == entryy4:
    pass 
else:
    messagebox.showinfo('')
 
def proceed():
    fn = entry1.get()
    ln = entry2.get()
    em = entry3.get()
    pas = entry4.get()
    if not fn.strip() or not ln.strip() or not em.strip() or not pas.strip():
        messagebox.showerror("HAaaA", "check mo ulit ")
    else:
        newWindow = Toplevel()
        newWindow.geometry("300x300")
        newWindow.title("New Window")

        text = Label(newWindow, text="new window").pack()
        newWindow.mainloop()

buttonF = Button(window, text="Facilitator", command=proceed)
buttonF.pack()
buttonS = Button(window, text="Student", command=proceed)
buttonS.pack()


window.mainloop()