import os
from tkinter import *
from openpyxl import Workbook, load_workbook

# <-----------------------------------------------------openpyxl----------------------------------------------->
    
fileName = "Infos.xlsx"

def createExcel():
    if not os.path.exists(fileName):
        wb = Workbook()
        ws = wb.active
        ws.append(["First Name", "Last Name", "Email", "Contact No.", "Nationality","Religion","Sex", "Civil Status", "Age", "Disability", "Permanent Address","Password"])
        wb.save(fileName)

        
def addInfo():
    Fname = firstName.get()
    Lname = lastName.get()
    Email= email.get()
    contact = contactNo.get()   
    nation = nationality.get() 
    

    wb = load_workbook(fileName)
    ws = wb.active
    ws.append([Fname, Lname, Email, contact, nation,"Religion","Sex", "Civil Status", "Age", "Disability", "Permanent Address","Password"])
    wb.save(fileName)


# <-----------------------------------------------------TKINTER------------------------------------------------->

mainWindow = Tk()
mainWindow.title("Try")
mainWindow.geometry("500x700")
mainWindow.resizable(False, False)

text1  = Label(mainWindow, text="first name")
text1.pack()

firstName= Entry(mainWindow)
firstName.pack()


text2  = Label(mainWindow, text="last name")
text2.pack()

lastName= Entry(mainWindow)
lastName.pack()


text3  = Label(mainWindow, text="email")
text3.pack()

email= Entry(mainWindow)
email.pack()

text4  = Label(mainWindow, text="contact")
text4.pack()

contactNo= Entry(mainWindow)
contactNo.pack()

text5  = Label(mainWindow, text="nationality")
text5.pack()

nationality= Entry(mainWindow)
nationality.pack()

text6  = Label(mainWindow, text="religion")
text6.pack()

religion= Entry(mainWindow)
religion.pack()

text7  = Label(mainWindow, text="sex")
text7.pack()

sex= Entry(mainWindow)
sex.pack()

text8  = Label(mainWindow, text="civil status")
text8.pack()

civilStatus= Entry(mainWindow)
civilStatus.pack()

text9  = Label(mainWindow, text="nationality")
text9.pack()

age= Entry(mainWindow)
age.pack()

text10  = Label(mainWindow, text="nationality")
text10.pack()

disability= Entry(mainWindow)
disability.pack()

text11  = Label(mainWindow, text="nationality")
text11.pack()

permanentAdd= Entry(mainWindow)
permanentAdd.pack()

text12  = Label(mainWindow, text="nationality")
text12.pack()

password= Entry(mainWindow)
password.pack()

saveBtn = Button(mainWindow, text="Save", command=addInfo)
saveBtn.pack()

createExcel()
mainWindow.mainloop()