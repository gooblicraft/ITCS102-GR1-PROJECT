import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

# Step 1: Create Excel file with users
file_path = "users.xlsx"
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Credentials"
    ws.append(["Username", "Password"])
    ws.append(["admin", "admin123"])
    ws.append(["user1", "pass1"])
    ws.append(["user2", "pass2"])
    wb.save(file_path)

# Step 2: Load credentials from Excel
def load_credentials(path):
    wb = load_workbook(path)
    ws = wb.active
    return {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True)}

credentials = load_credentials(file_path)

# Step 3: Open next window
def open_next_window():
    next_win = tk.Toplevel()
    next_win.title("Welcome")
    tk.Label(next_win, text="Login Successful!", font=("Helvetica", 16)).pack(pady=20)
    next_win.geometry("300x100")

# Step 4: Validate login
def validate_login():
    username = entry_username.get()
    password = entry_password.get()
    if credentials.get(username) == password:
        messagebox.showinfo("Login", "Login successful!")
        root.destroy()
        main_app = tk.Tk()
        main_app.withdraw()
        open_next_window()
        main_app.mainloop()
    else:
        messagebox.showerror("Login", "Invalid username or password.")

# Step 5: Build login window
root = tk.Tk()
root.title("Login")
root.geometry("300x180")

tk.Label(root, text="Username").pack(pady=5)
entry_username = tk.Entry(root)
entry_username.pack()

tk.Label(root, text="Password").pack(pady=5)
entry_password = tk.Entry(root, show="*")
entry_password.pack()

tk.Button(root, text="Login", command=validate_login).pack(pady=15)

root.mainloop()
