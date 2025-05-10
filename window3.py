from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
import openpyxl
import os, webbrowser, subprocess
import os
import datetime
from openpyxl import Workbook, load_workbook
from tkinter import messagebox
import cv2
import qrcode
from pyzbar.pyzbar import decode
from PIL import Image, ImageTk
import threading
import random
import time

def logOut():
    window.destroy()
    subprocess.Popen(['python', 'window1.py'])

def open_github():
    webbrowser.open("https://github.com/gooblicraft/ITCS102-GR1-PROJECT")
    
def open_readme():
    webbrowser.open("https://github.com/gooblicraft/ITCS102-GR1-PROJECT?tab=readme-ov-file#itcs102-gr1-project")

def save_account_data(account_id):
    try:
        wb = load_workbook("AccountDatabase.xlsx")
        ws = wb.active

        # Find the header row to map column indices
        headers = [cell.value for cell in ws[1]]

        # Build a dict of updated values from the form
        updated_data = {
            "First Name": firstName_entry.get(),
            "Last Name": lastName_entry.get(),
            "Email": email_entry.get(),
            "Contact Number": contact_number_entry.get(),
            "Nationality": nationality_entry.get(),
            "Religion": religion_combobox.get(),
            "Sex": sex_combobox.get(),
            "Civil Status": civil_status_combobox.get(),
            "Age": age_entry.get(),
            "Disability": isDisability_entry.get(),
            "Permanent Address": permanent_address_entry.get()
        }

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value).strip() == account_id:
                for cell in row:
                    header = headers[cell.column - 1]
                    if header in updated_data:
                        cell.value = updated_data[header]
                wb.save("AccountDatabase.xlsx")
                print("Account data updated successfully.")
                
                # ✅ Update full name label
                firstName_label.configure(text=updated_data["First Name"])
                lastName_label.configure(text=updated_data["Last Name"])
                update_full_name_label()
                
                return
        print("Account ID not found.")

    except FileNotFoundError:
        print("ERROR: AccountDatabase.xlsx not found.")

def update_full_name_label():
    full_name = f"{firstName_label.cget('text')} {lastName_label.cget('text')}"
    fullName_label.configure(text=full_name)


def toggle_show():
    toggle_button_image.configure(file="assets\\window3\\account_tab\\image_btn_edit.png")
    # Get text from entry and update label *before* showing it
    firstName_label.configure(text=firstName_entry.get())
    lastName_label.configure(text=lastName_entry.get())
    nationality_label.configure(text=nationality_entry.get())
    religion_label.configure(text=religion_combobox.get())
    age_label.configure(text=age_entry.get())
    sex_label.configure(text=sex_combobox.get())
    civil_status_label.configure(text=civil_status_combobox.get())
    isDisability_label.configure(text=isDisability_entry.get())
    contact_number_label.configure(text=contact_number_entry.get())
    email_label.configure(text=email_entry.get())
    permanent_address_label.configure(text=permanent_address_entry.get())

    # Show label
    canvas.itemconfigure(firstName_label_canvas, state='normal')
    canvas.itemconfigure(lastName_label_canvas, state='normal')
    canvas.itemconfigure(nationality_label_canvas, state='normal')
    canvas.itemconfigure(religion_label_canvas, state='normal')
    canvas.itemconfigure(age_label_canvas, state='normal')
    canvas.itemconfigure(sex_label_canvas, state='normal')
    canvas.itemconfigure(civil_status_label_canvas, state='normal')
    canvas.itemconfigure(isDisability_label_canvas, state='normal')
    canvas.itemconfigure(contact_number_label_canvas, state='normal')
    canvas.itemconfigure(email_label_canvas, state='normal')
    canvas.itemconfigure(permanent_address_label_canvas, state='normal')
    
    # Hide entry
    canvas.itemconfigure(firstName_entry_canvas, state='hidden')
    canvas.itemconfigure(lastName_entry_canvas, state='hidden')
    canvas.itemconfigure(nationality_entry_canvas, state='hidden')
    canvas.itemconfigure(religion_combobox_canvas, state='hidden')
    canvas.itemconfigure(age_entry_canvas, state='hidden')
    canvas.itemconfigure(sex_combobox_canvas, state='hidden')
    canvas.itemconfigure(civil_status_combobox_canvas, state='hidden')
    canvas.itemconfigure(isDisability_entry_canvas, state='hidden')
    canvas.itemconfigure(contact_number_entry_canvas, state='hidden')
    canvas.itemconfigure(email_entry_canvas, state='hidden')
    canvas.itemconfigure(permanent_address_entry_canvas, state='hidden')
    
    toggle_button.configure(text="Edit", command=toggle_edit)

def toggle_edit():
    toggle_button_image.configure(file="assets\\window3\\account_tab\\image_btn_show.png")
    # Optionally: prefill entry with label text
    firstName_entry.delete(0, 'end')
    firstName_entry.insert(0, firstName_label.cget("text"))
    
    lastName_entry.delete(0, 'end')
    lastName_entry.insert(0, lastName_label.cget("text"))
    
    nationality_entry.delete(0, 'end')
    nationality_entry.insert(0, nationality_label.cget("text"))
    
    religion_combobox.delete(0, 'end')
    religion_combobox.insert(0, religion_label.cget("text"))
    
    age_entry.delete(0, 'end')
    age_entry.insert(0, age_label.cget("text"))
    
    sex_combobox.delete(0, 'end')
    sex_combobox.insert(0, sex_label.cget("text"))
    
    civil_status_combobox.delete(0, 'end')
    civil_status_combobox.insert(0, civil_status_label.cget("text"))
    
    isDisability_entry.delete(0, 'end')
    isDisability_entry.insert(0, isDisability_label.cget("text"))
    
    contact_number_entry.delete(0, 'end')
    contact_number_entry.insert(0, contact_number_label.cget("text"))
    
    email_entry.delete(0, 'end')
    email_entry.insert(0, email_label.cget("text"))
    
    permanent_address_entry.delete(0, 'end')
    permanent_address_entry.insert(0, permanent_address_label.cget("text"))
    
    # Hide label
    canvas.itemconfigure(firstName_label_canvas, state='hidden')
    canvas.itemconfigure(lastName_label_canvas, state='hidden')
    canvas.itemconfigure(nationality_label_canvas, state='hidden')
    canvas.itemconfigure(religion_label_canvas, state='hidden')
    canvas.itemconfigure(age_label_canvas, state='hidden')
    canvas.itemconfigure(sex_label_canvas, state='hidden')
    canvas.itemconfigure(civil_status_label_canvas, state='hidden')
    canvas.itemconfigure(isDisability_label_canvas, state='hidden')
    canvas.itemconfigure(contact_number_label_canvas, state='hidden')
    canvas.itemconfigure(email_label_canvas, state='hidden')
    canvas.itemconfigure(permanent_address_label_canvas, state='hidden')
    
    # Show entry
    canvas.itemconfigure(firstName_entry_canvas, state='normal')
    canvas.itemconfigure(lastName_entry_canvas, state='normal')
    canvas.itemconfigure(nationality_entry_canvas, state='normal')
    canvas.itemconfigure(religion_combobox_canvas, state='normal')
    canvas.itemconfigure(age_entry_canvas, state='normal')
    canvas.itemconfigure(sex_combobox_canvas, state='normal')
    canvas.itemconfigure(civil_status_combobox_canvas, state='normal')
    canvas.itemconfigure(isDisability_entry_canvas, state='normal')
    canvas.itemconfigure(contact_number_entry_canvas, state='normal')
    canvas.itemconfigure(email_entry_canvas, state='normal')
    canvas.itemconfigure(permanent_address_entry_canvas, state='normal')
    
    toggle_button.configure(text="Show", command=toggle_show)
    
accountID = None
accountType = None
firstName = None
lastName = None
email = None
contact_number = None
nationality = None
religion = None
sex = None
civil_status = None
age = None
isDisability = None
permanent_address = None
password = None

def load_account_data(account_id):
    global accountID, accountType, firstName, lastName, email, contact_number, nationality, religion, sex, civil_status, age, isDisability, permanent_address, password

    try:
        wb = load_workbook("AccountDatabase.xlsx")
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            current_id = str(row[0]).strip()
            if current_id == account_id:
                print(f"Details for Account ID: {account_id}")
                for header, value in zip(headers, row):
                    
                    if header == "ID Number":
                        accountID = value
                        print(f"Account ID: {accountID}")
                    elif header == "Account Type":
                        accountType = value
                        print(f"Account Type: {accountType}")
                    elif header == "First Name":
                        firstName = value
                        print(f"Account First Name: {firstName}")
                    elif header == "Last Name":
                        lastName = value
                        print(f"Account Last Name: {lastName}")
                    elif header == "Email":
                        email = value
                        print(f"Account Email: {email}")
                    elif header == "Contact Number":
                        contact_number = value
                        print(f"Contact Number: {contact_number}")
                    elif header == "Nationality":
                        nationality = value
                        print(f"Nationality : {nationality}")
                    elif header == "Religion":
                        religion = value
                        print(f"Religion : {religion}")
                    elif header == "Sex":
                        sex = value
                        print(f"Sex : {sex}")
                    elif header == "Civil Status":
                        civil_status = value
                        print(f"Civil Status : {civil_status}")
                    elif header == "Age":
                        age = value
                        print(f"Age : {age}")
                    elif header == "Disability":
                        isDisability = value
                        print(f"Disabillity : {isDisability}")
                    elif header == "Permanent Address":
                        permanent_address = value
                        print(f"Permanent Address : {permanent_address}")
                    elif header == "Password":
                        password = value
                        print(f"Password : {password}")
                    
                    # print(f"{header}: {value}")
                return accountID, accountType, firstName, lastName, email, contact_number, nationality, religion, sex, civil_status, age, isDisability, permanent_address, password

        print(f"Account ID {account_id} not found.")

    except FileNotFoundError:
        print("ERROR: AccountDatabase.xlsx not found.")

# Nakuha nayung account id, itutugma nalang sa account database then ishoshow dito yung code sa tab1
account_id = os.environ.get("ACCOUNT_ID")

if not account_id:
    print("ERROR: ACCOUNT_ID environment variable not set.")
else:
    load_account_data(account_id)
    print(f"Received Account ID: {account_id}")
    print(accountID, accountType, lastName, email, contact_number, nationality, religion, sex, civil_status, age, isDisability, permanent_address, password)

window = Tk()
window.geometry("665x640")
window.iconbitmap("assets\\logo.ico")
window.configure(bg="#FFFFFF")

# =========== SECTION FOR STYLING TTK ==============
style = ttk.Style()
style.theme_use('default')
bold_font = ("JetBrains Mono", 10, "bold")
style.configure("TNotebook", background="#FFFFFF", borderwidth=0)
style.configure("TNotebook.Tab", background="#FFFFFF", foreground="#000000", padding=10, font=bold_font)
style.configure("Custom.TCombobox",
    foreground="#9F26C7",
    fieldbackground="#EEE9E9",
    background="#EEE9E9",
    bordercolor="#EEE9E9",
    lightcolor="#EEE9E9",
    darkcolor="#EEE9E9",
    highlightthickness=0,
    borderwidth=0,
    relief="flat"
)

# 💡 Prevent theme from overriding colors on selection/focus
style.map("Custom.TCombobox",
    foreground=[("active", "#9F26C7"), ("!disabled", "#9F26C7")],
    fieldbackground=[("readonly", "#EEE9E9"), ("!disabled", "#EEE9E9"), ("active", "#EEE9E9")],
    background=[("readonly", "#EEE9E9"), ("!disabled", "#EEE9E9"), ("active", "#EEE9E9")],
    highlightcolor=[("focus", "#EEE9E9")],
    bordercolor=[("focus", "#EEE9E9")]
)
style.layout("TNotebook.Tab", [('Notebook.tab', {'sticky': 'nswe', 'children': [('Notebook.padding', {'side': 'top', 'sticky': 'nswe', 'children': [('Notebook.label', {'side': 'top', 'sticky': ''})],})],})])
style.map("TNotebook.Tab", background=[("selected", "#0E3269")], foreground=[("selected", "#FFFFFF")])

# ============ SECTION FOR TABS ============
notebook = ttk.Notebook(window, style='TNotebook')
tab1 = Frame(notebook, bg="#FFFFFF", width=600, height=500)
tab3 = Frame(notebook, bg="#0E3269")  # QR Code tab
notebook.add(tab1, text="Account")
notebook.add(tab3, text="QR Code")

# Add tabs conditionally
if accountType and accountType.lower() == "facilitator":
    tab2 = Frame(notebook, bg="#0E3269", width=600, height=460)
    tab4 = Frame(notebook, bg="#0E3269", width=600, height=460)
    notebook.add(tab2, text="QR Scan")
    notebook.add(tab4, text="Attendance")

notebook.place(x=32, y=120)

# ==================== SECTION FOR TAB 1 (ACCOUNT SUMMARY TAB) =====================================
# DITO ED :>

canvas = Canvas(tab1, bg="#FFFFFF", height=460, width=600, bd=0, highlightthickness=0, relief="ridge")
canvas.place(x=0, y=0)

# ============ SECTION FOR IMAGES ============
image_image_1 = PhotoImage(file="assets/window3/account_tab/image_1.png")
image_1 = Label(window, image=image_image_1, bg="white")
image_1.place(x=-2, y=2)

image_image_2 = PhotoImage(file="assets/window3/account_tab/image_2.png")
image_2 = canvas.create_image(300, 230, image=image_image_2)

# ============ SECTION FOR IMAGE BUTTONS ============
button_image_1 = PhotoImage(file="assets/window3/account_tab/button_1.png")
button_about = Button(window, image=button_image_1, borderwidth=0, highlightthickness=0, command=open_readme, relief="flat")
button_about.place(x=346.0, y=14.0, width=82.0, height=30.0)

button_image_2 = PhotoImage(file="assets/window3/account_tab/button_2.png")
button_github = Button(window, image=button_image_2, borderwidth=0, highlightthickness=0, command=open_github, relief="flat")
button_github.place(x=430.0, y=14.0, width=82.0, height=30.0)

button_image_3 = PhotoImage(file="assets/window3/account_tab/button_3.png")
button_logOut = Button(window, image=button_image_3, borderwidth=0, highlightthickness=0, command=logOut, relief="flat")
button_logOut.place(x=514.0, y=14.0, width=82.0, height=30.0)

toggle_button_image = PhotoImage(file="assets\\window3\\account_tab\\image_btn_edit.png")
toggle_button = Button(tab1, image=toggle_button_image, borderwidth=0, border=0, highlightthickness=0, relief="flat", command=toggle_edit)
toggle_button.place(x=380, y=420)

save_button_image = PhotoImage(file="assets\\window3\\account_tab\\image_btn_save.png")
save_button = Button(tab1, image=save_button_image, borderwidth=0, border=0, highlightthickness=0, relief="flat", command=lambda: save_account_data(account_id))
save_button.place(x=490, y=420)

# ============ SECTION FOR NEEDED WIDGET ============
firstName_label = Label(tab1,bd=0,text=firstName,bg="#EEE9E9",fg="#0E3269",font= ("JetBrains Mono", 10),highlightthickness=0,width=23,anchor="w")
firstName_label_canvas = canvas.create_window(262, 94, window=firstName_label, state="normal")

firstName_entry = Entry(tab1,bd=0,bg="#EEE9E9",fg="#9F26C7",font= ("JetBrains Mono", 10),highlightthickness=0,width=23)
firstName_entry_canvas = canvas.create_window(262, 94, window=firstName_entry, state="hidden")

lastName_label = Label(tab1,bd=0,text=lastName,bg="#EEE9E9",fg="#0E3269",font= ("JetBrains Mono", 10),highlightthickness=0,width=23,anchor="w")
lastName_label_canvas = canvas.create_window(479, 94, window=lastName_label, state="normal")

lastName_entry = Entry(tab1,bd=0,bg="#EEE9E9",fg="#9F26C7",font= ("JetBrains Mono", 10),highlightthickness=0, width=23)
lastName_entry_canvas = canvas.create_window(479, 94, window=lastName_entry, state="hidden")

nationality_label = Label(tab1,bd=0,text=nationality,bg="#EEE9E9",fg="#0E3269",font= ("JetBrains Mono", 10),highlightthickness=0,width=16,anchor="w")
nationality_label_canvas = canvas.create_window(234, 159, window=nationality_label, state="normal")

nationality_entry = Entry(tab1,bd=0,bg="#EEE9E9",fg="#9F26C7",font= ("JetBrains Mono", 10),highlightthickness=0,width=16)
nationality_entry_canvas = canvas.create_window(234, 159, window=nationality_entry, state="hidden")

religion_label = Label(tab1,bd=0,text=religion,bg="#EEE9E9",fg="#0E3269",font= ("JetBrains Mono", 10),highlightthickness=0,width=13,anchor="w")
religion_label_canvas = canvas.create_window(392, 159, window=religion_label, state="normal")

# Combobox religion
religion_combobox = ttk.Combobox(values=["Catholic", "INC", "Muslim"], style='Custom.TCombobox', font=("JetBrains Mono", 10),foreground="#9F26C7",width=12, state="readonly")
religion_combobox.set(religion)
religion_combobox_canvas = canvas.create_window(392, 159, window=religion_combobox, state='hidden')

age_label = Label(tab1,bd=0,text=age,bg="#EEE9E9",fg="#0E3269",font= ("JetBrains Mono", 10),highlightthickness=0,width=8,anchor="w")
age_label_canvas = canvas.create_window(522, 159, window=age_label, state="normal")

age_entry = Entry(tab1,bd=0,bg="#EEE9E9",fg="#9F26C7",font= ("JetBrains Mono", 10),highlightthickness=0,width=8)
age_entry_canvas = canvas.create_window(522, 159, window=age_entry, state="hidden")

sex_label = Label(tab1,bd=0,text=sex,bg="#EEE9E9",fg="#0E3269",font= ("JetBrains Mono", 10),highlightthickness=0,width=8,anchor="w")
sex_label_canvas = canvas.create_window(202, 232, window=sex_label, state="normal")

sex_combobox = ttk.Combobox(values=["Male", "Female"], style='Custom.TCombobox', font=("JetBrains Mono", 10),width=7, state="readonly")
sex_combobox.set(sex)
sex_combobox_canvas = canvas.create_window(202, 232, window=sex_combobox, state='hidden')

civil_status_label = Label(tab1,bd=0,text=civil_status,bg="#EEE9E9",fg="#0E3269",font= ("JetBrains Mono", 10),highlightthickness=0,width=12,anchor="w")
civil_status_label_canvas = canvas.create_window(342, 232, window=civil_status_label, state="normal")

civil_status_combobox = ttk.Combobox(values=["Single", "Married"], style='Custom.TCombobox', font=("JetBrains Mono", 10),width=11, state="readonly")
civil_status_combobox.set(civil_status)
civil_status_combobox_canvas = canvas.create_window(345, 232, window=civil_status_combobox, state='hidden')

isDisability_label = Label(
    tab1,
    bd=0,
    text=isDisability,
    bg="#EEE9E9",
    fg="#0E3269",
    font= ("JetBrains Mono", 10),
    highlightthickness=0,
    width=6,
    anchor="w"
)
isDisability_label_canvas = canvas.create_window(467, 232, window=isDisability_label, state="normal")

isDisability_entry = Entry(
    tab1,
    bd=0,
    bg="#EEE9E9",
    fg="#9F26C7",
    font= ("JetBrains Mono", 10),
    highlightthickness=0,
    width=6
)
isDisability_entry_canvas = canvas.create_window(467, 232, window=isDisability_entry, state="hidden")

contact_number_label = Label(
    tab1,
    bd=0,
    text=contact_number,
    bg="#EEE9E9",
    fg="#0E3269",
    font= ("JetBrains Mono", 12),
    highlightthickness=0,
    width=15,
    anchor="w"
)
contact_number_label_canvas = canvas.create_window(247, 372, window=contact_number_label, state="normal")

contact_number_entry = Entry(
    tab1,
    bd=0,
    bg="#EEE9E9",
    fg="#9F26C7",
    font= ("JetBrains Mono", 12),
    highlightthickness=0,
    width=15
)
contact_number_entry_canvas = canvas.create_window(247, 373, window=contact_number_entry, state="hidden")

email_label = Label(
    tab1,
    bd=0,
    text=email,
    bg="#EEE9E9",
    fg="#0E3269",
    font= ("JetBrains Mono", 10),
    highlightthickness=0,
    width=23,
    anchor="w"
)
email_label_canvas = canvas.create_window(485, 372, window=email_label, state="normal")

email_entry = Entry(
    tab1,
    bd=0,
    bg="#EEE9E9",
    fg="#9F26C7",
    font= ("JetBrains Mono", 10),
    highlightthickness=0,
    width=23
)
email_entry_canvas = canvas.create_window(485, 372, window=email_entry, state="hidden")

permanent_address_label = Label(
    tab1,
    bd=0,
    text=permanent_address,
    bg="#EEE9E9",
    fg="#0E3269",
    font= ("JetBrains Mono", 10),
    highlightthickness=0,
    width=23,
    anchor="w"
)
permanent_address_label_canvas = canvas.create_window(262, 442, window=permanent_address_label, state="normal")

permanent_address_entry = Entry(
    tab1,
    bd=0,
    bg="#EEE9E9",
    fg="#9F26C7",
    font= ("JetBrains Mono", 10),
    highlightthickness=0,
    width=23
)
permanent_address_entry_canvas = canvas.create_window(262, 442, window=permanent_address_entry, state="hidden")

accountID_label = Label(
    tab1,
    bd=0,
    text=account_id,
    bg="#9F26C7",
    fg="#FFFFFF",
    font= ("JetBrains Mono", 8),
    highlightthickness=0,
    width=10
)
canvas.create_window(66, 162, window=accountID_label, state="normal")

account_type_label = Label(
    tab1,
    bd=0,
    text=accountType,
    bg="#FFFFFF",
    fg="#767676",
    font= ("Lalezar", 10),
    highlightthickness=0,
    width=12
)
canvas.create_window(66, 187, window=account_type_label, state="normal")

full_name = f"{firstName_label.cget("text")} {lastName_label.cget("text")}"
fullName_label = Label(
    tab1,
    bd=0,
    text=full_name,
    bg="#0E3269",
    fg="#FFFFFF",
    font= ("JetBrains Mono", 8),
    highlightthickness=0,
    width=16
)
canvas.create_window(66, 137, window=fullName_label, state="normal")


# MOST COMPLICATED RAAAAAAAAAAAAAAAAAHHHHHHHHHHH

attendance_filename = "AttendanceDatabase.xlsx"
clock_in_status = {} 
total_time_ins = 0  

def createAttendanceExcel():
    if not os.path.exists(attendance_filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Account ID", "First Name", "Last Name", "Sex", "Age", "Date", "Time In", "Time Out", "Status"])
        workbook.save(attendance_filename)

def time_in(account_id, first_name, last_name, sex, age, timestamp):
    global total_time_ins
    
    if account_id in clock_in_status:
        last_time_in = clock_in_status[account_id]["time_in"]
        current_time = datetime.now()

        if current_time - last_time_in < timedelta(hours=1):
            time_left = timedelta(hours=1) - (current_time - last_time_in)
            messagebox.showerror("Cooldown Error", f"You can only clock in again after {time_left}.")
            return

    clock_in_status[account_id] = {"status": "Present", "time_in": timestamp}
    total_time_ins += 1  # Increase the total time-in count

    log_attendance(account_id, first_name, last_name, sex, age, "Present", timestamp, None)
    messagebox.showinfo("Successfully Timed In", f"{first_name} {last_name} has timed in at {timestamp}")

def time_out(account_id, first_name, last_name, sex, age, timestamp):
    time_in_time = clock_in_status.get(account_id, {}).get("time_in", None)
    if not time_in_time:
        messagebox.showerror("Error", f"{first_name} {last_name} was not time-in.")
        return

    log_attendance(account_id, first_name, last_name, sex, age, "Present", time_in_time, timestamp)
    
    del clock_in_status[account_id]  

    messagebox.showinfo("Time Out", f"{first_name} {last_name} has timed out at {timestamp}")

def log_attendance(account_id, first_name, last_name, sex, age, status, time_in_time, time_out_time):
    if not os.path.exists(attendance_filename):
        createAttendanceExcel()

    workbook = load_workbook(attendance_filename)
    sheet = workbook.active

    row_to_update = None
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == account_id and row[8].value == "Present":  
            row_to_update = row
            break

    if row_to_update:
        row_to_update[7].value = time_out_time.split(" ")[1][:5] if time_out_time else "" 
    else:
        sheet.append([account_id, first_name, last_name, sex, age, time_in_time.split(" ")[0], time_in_time.split(" ")[1][:5], time_out_time.split(" ")[1][:5] if time_out_time else "", "Present"])

    workbook.save(attendance_filename)

    write_present_count(sheet)

def write_present_count(sheet):
    global total_time_ins
    
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == "Total Present":
            sheet.delete_rows(row[0].row)

    sheet.append(["Total Present", total_time_ins, "", "", "", "", "", ""])

    sheet.parent.save(attendance_filename)

def process_qr_data(data):
    try:
        user_data = data.split("\n")
        user_info = {item.split(":")[0].strip(): item.split(":")[1].strip() for item in user_data}

        account_id = user_info.get("ID Number")
        first_name = user_info.get("First Name")
        last_name = user_info.get("Last Name")
        sex = user_info.get("Sex")
        age = user_info.get("Age")
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        if not account_id:
            messagebox.showerror("Error", "QR Code is missing Account ID.")
            return

     
        if account_id in clock_in_status:
            time_out(account_id, first_name, last_name, sex, age, timestamp)
        else:
            time_in(account_id, first_name, last_name, sex, age, timestamp)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to process QR code: {e}")


video_frame = Frame(tab2, width=320, height=240)
video_frame.place(relx=0.5, rely=0.5, anchor="center") 

video_label = Label(video_frame) 
video_label.pack(fill="both", expand=True)

cap = None

def start_scanning():
    global cap
    if cap is None or not cap.isOpened():
        cap = cv2.VideoCapture(0)  
        if not cap.isOpened():
            messagebox.showerror("Error", "Failed to open the webcam.")
            return
        update_frame()  

def update_frame():
    if cap is None or not cap.isOpened():
        return

    ret, frame = cap.read()
    if not ret:
        messagebox.showerror("Error", "Failed to capture image from webcam")
        return

    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

    img = Image.fromarray(frame_rgb)
    img = img.resize((320, 240))  
    img_tk = ImageTk.PhotoImage(img)


    video_label.img_tk = img_tk  
    video_label.config(image=img_tk)

    decoded_objects = decode(frame)
    for obj in decoded_objects:
        qr_data = obj.data.decode('utf-8')
        qr_type = obj.type
        if qr_type == "QRCODE":
           
            process_qr_data(qr_data)
            break  

    tab2.after(20, update_frame) 

start_scanning()

pic = Frame(tab3, width=300, height=300)
pic.place(relx=0.5, rely=0.5, anchor="center")

qr_image = Image.open("assets/window3/account_tab/qr_image.png")
qr_image_resized = qr_image.resize((300, 300))  # Resize to 100x100 pixels

qr_image_tk = ImageTk.PhotoImage(qr_image_resized)

qr_img = Label(pic, image=qr_image_tk)
qr_img.pack()

qr_img.image = qr_image_tk


#==================== SECTION FOR TAB 3 (QR SHOW TAB) =====================================


# def show_user_data(account_id):
#     path = "AccountDatabase.xlsx"
#     workbook = load_workbook(path)
#     sheet = workbook.active
    
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         if row[0] == account_id:  
#             first_name, last_name, email, contact_num, nationality, religion, sex, civil_status, age, disability, address, password = row[2:14]
#             break

#     user_data = f"ID: {account_id}\nName: {first_name} {last_name}\nEmail: {email}\nContact: {contact_num}\nNationality: {nationality}\nReligion: {religion}\nSex: {sex}\nCivil Status: {civil_status}\nAge: {age}\nDisability: {disability}\nAddress: {address}"

#     qr_path = generate_qr_code(user_data, last_name)


#     # Display user data
#     user_info_label = Label(tab3, text=user_data, justify="left", font=("Arial", 10))
#     user_info_label.pack(pady=20)

#     if qr_path:
#         img = Image.open(qr_path)
#         img = img.resize((150, 150))  # Resize image to fit the window
#         photo = ImageTk.PhotoImage(img)

#         qr_label = Label(tab3, image=photo)
#         qr_label.image = photo 
#         qr_label.pack(pady=20)


# ==================== SECTION FOR TAB 4 (ATTENDANCE TAB) =====================================
# DITO KYLAA :>

window.resizable(False, False)
window.mainloop()
