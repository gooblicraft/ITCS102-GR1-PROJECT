import subprocess
from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook
import os

class LoginWindow:
    def __init__(self):
        self.window = Tk()
        self.window.geometry("665x410")
        self.window.iconbitmap("assets\\logo.ico")
        self.window.resizable(False,False)
        self.window.configure(bg="#FFFFFF")
        self.window.title("Window 1")
        self.password_visible = False
        self.account_id = None

        self.setup_ui()
        self.center_window(665, 410)
        self.window.bind("<Configure>", self.on_configure)
        self.window.mainloop()

    def center_window(self, width, height):
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.window.geometry(f"{width}x{height}+{x}+{y}")

    def on_configure(self, event):
        if self.window.state() == 'zoomed':
            self.window.state('normal')
            self.center_window(700, 600)

    def setup_ui(self):
        canvas = Canvas(self.window, bg="#FFFFFF", height=410, width=665, bd=0, highlightthickness=0, relief="ridge")
        canvas.place(x=0, y=0)

        self.image_1 = PhotoImage(file="assets\\window1\\image_1.png")
        canvas.create_image(519.0, 126.0, image=self.image_1)

        self.image_2 = PhotoImage(file="assets\\window1\\image_2.png")
        canvas.create_image(519.0, 190.0, image=self.image_2)

        self.image_3 = PhotoImage(file="assets\\window1(new)\\image (2).png")
        canvas.create_image(170.0, 205.0, image=self.image_3)

        canvas.create_text(466.0, 76.0, anchor="nw", text="Sign in your account",
                           fill="#757575", font=("JetBrains Mono", 10 * -1))
        canvas.create_text(419.0, 38.0, anchor="nw", text="ScanTracker",
                            fill="#060606", font=("JetBrains Mono", 24, "bold"))

        self.username_entry = Entry(bd=0, bg="#AEAEAE", fg="#767676",
                                    highlightthickness=0, font=("JetBrains Mono", 10, "bold"))
        self.username_entry.insert(0, "Username")
        self.username_entry.place(x=464.0, y=118.0, width=152.0, height=14.0)
        self.username_entry.bind("<FocusIn>", lambda e: self.on_entry_click(self.username_entry, "Username"))
        self.username_entry.bind("<FocusOut>", lambda e: self.on_focusout(self.username_entry, "Username"))

        self.password_entry = Entry(bd=0, bg="#AEAEAE", fg="#767676",
                                    highlightthickness=0, font=("JetBrains Mono", 10, "bold"))
        self.password_entry.insert(0, "Password")
        self.password_entry.place(x=464.0, y=182.0, width=152.0, height=14.0)
        self.password_entry.bind("<FocusIn>", lambda e: self.on_entry_click(self.password_entry, "Password", True))
        self.password_entry.bind("<FocusOut>", lambda e: self.on_focusout(self.password_entry, "Password", True))

        self.button_image_1 = PhotoImage(file="assets\\window1(new)\\Login.png")
        Button(image=self.button_image_1, borderwidth=0, highlightthickness=0,
            command=self.validate_login, relief="flat", cursor="hand2").place(
            x=402.0, y=224.0, width=234.0, height=59.0)

        self.show_button = Button(self.window, text="Show", command=self.toggle_password_button,
                                font=("JetBrains Mono", 7), bd=0, bg="#AEAEAE",
                                activebackground="#AEAEAE", fg="gray25", cursor="hand2")
        self.show_button.place(x=568, y=180, width=40, height=18)

        self.button_image_2 = PhotoImage(file="assets\\window1(new)\\create.png")
        Button(image=self.button_image_2, borderwidth=0, highlightthickness=0,
            command=self.open_window2, relief="flat", cursor="hand2").place(
            x=402.0, y=325.0, width=234.0, height=57.0)

        # ✅ Forgot Password clickable label
        forgot_label = Label(self.window, text="Forgot Password?", fg="#9F26C7", cursor="hand2",
                            bg="#FFFFFF", font=("JetBrains Mono", 8, "underline"))
        forgot_label.place(x=495, y=284)
        forgot_label.bind("<Button-1>", lambda e: self.forgot_password())

    def toggle_password_button(self):
        if self.password_entry.get() != "Password":
            if self.password_visible:
                self.password_entry.config(show="*")
                self.show_button.config(text="Show")
            else:
                self.password_entry.config(show="")
                self.show_button.config(text="Hide")
            self.password_visible = not self.password_visible

    def on_entry_click(self, entry, placeholder, is_password=False):
        if entry.get() == placeholder:
            entry.delete(0, "end")
            entry.config(fg="gray33")
            if is_password:
                entry.config(show="" if self.password_visible else "*")

    def on_focusout(self, entry, placeholder, is_password=False):
        if entry.get() == "":
            entry.insert(0, placeholder)
            entry.config(fg="#767676")
            if is_password:
                entry.config(show="")

    def validate_login(self):
        input_username = self.username_entry.get().strip().lower()
        input_password = self.password_entry.get().strip()

        try:
            wb = load_workbook("AccountDatabase.xlsx")
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                account_id = str(row[0]).strip() if row[0] else ""
                account_type = str(row[1]).strip() if row[1] else ""
                first_name = str(row[2]).strip().lower() if row[2] else ""
                last_name = str(row[3]).strip().lower() if row[3] else ""
                password = str(row[-1]).strip() if row[-1] else ""

                full_name = f"{first_name} {last_name}".strip()

                if (input_username == account_id or input_username == full_name) and input_password == password:
                    self.account_id = account_id
                    messagebox.showinfo("Login", f"Welcome, {account_type.capitalize()} {first_name.capitalize()}!")
                    self.logged_in()
                    self.window.destroy()
                    return

            messagebox.showerror("Login Failed", "Invalid username or password.")

        except FileNotFoundError:
            messagebox.showerror("File Error", "The file 'AccountDatabase.xlsx' was not found.")

    def open_window2(self):
        subprocess.Popen(['python', 'window2.py'])
        self.window.destroy()

    def logged_in(self):
        pack = os.environ.copy()
        if self.account_id:
            pack['ACCOUNT_ID'] = self.account_id
        subprocess.Popen(['python', 'window3.py'], env=pack)

    # ✅ FORGOT PASSWORD FEATURE
    def forgot_password(self):
        def submit_new_password():
            target_username = username_entry.get().strip().lower()
            new_pass = new_password_entry.get().strip()

            if not target_username or not new_pass:
                messagebox.showwarning("Input Error", "Please provide both username and new password.")
                return

            try:
                wb = load_workbook("AccountDatabase.xlsx")
                ws = wb.active

                updated = False
                for row in ws.iter_rows(min_row=2):
                    account_id = str(row[0].value).strip().lower() if row[0].value else ""
                    first_name = str(row[2].value).strip().lower() if row[2].value else ""
                    last_name = str(row[3].value).strip().lower() if row[3].value else ""
                    full_name = f"{first_name} {last_name}".strip()

                    if target_username == account_id or target_username == full_name:
                        row[-1].value = new_pass
                        updated = True
                        break

                if updated:
                    wb.save("AccountDatabase.xlsx")
                    messagebox.showinfo("Success", "Password reset successfully.")
                    reset_window.destroy()
                else:
                    messagebox.showerror("Error", "Username not found.")

            except FileNotFoundError:
                messagebox.showerror("Error", "AccountDatabase.xlsx not found.")

        reset_window = Toplevel(self.window)
        reset_window.title("Forgot Password")
        reset_window.geometry("300x180")
        reset_window.resizable(False, False)

        Label(reset_window, text="Account ID or Full Name:").pack(pady=5)
        username_entry = Entry(reset_window)
        username_entry.pack(pady=5)

        Label(reset_window, text="New Password:").pack(pady=5)
        new_password_entry = Entry(reset_window, show="*")
        new_password_entry.pack(pady=5)

        Button(reset_window, text="Submit", command=submit_new_password).pack(pady=10)

if __name__ == "__main__":
    LoginWindow()